#this program is meant to cycle through the ncbi gene expression pages and add them to an excel file
#and take RPKM values for the 27 tissue types.
#Justin Womack, Dec 2017
#Python3


#Program speed can be improved with pooling.  Limiting factor was not CPU
#but website ping request time


import openpyxl as op
import requests, re, json
from bs4 import BeautifulSoup as bs
from lxml import html


#loads workbook, gets active sheet
wb = op.load_workbook('data_match.xlsx')
sheet = wb.get_active_sheet()
max_rows = sheet.max_row
max_columns = sheet.max_column
error_list = []

for i in range(max_rows+1):

	if i < 24000:
		pass

	#allows for the program to be run up to a certain gene ID
	#mostly for testing.  Whole program will run to max value. probably
	#easier to run it until shee.cell.value == None
	#if want to run all at once I would remove the line below this
	#Take the if and else loop inside it, back tab it.  Make it
	#elif, and else.   The for loop should end it at the right place.
	
	elif i < 25374:
		#This was a control but it turns out its easier to run code
		#in chunks with different chunks running at a time.
		#Use the two above if/elif to get values for chunk you want

		#allows for program to skip over all values that has a name in column 32
		#towards the end of the list.  Since I added new labels the list got extended
		#this will rehit all the previous values
		if sheet.cell(row=i+1, column=32).value != None:
			print(i+1, 'skip', sheet.cell(row=i+1, column=3).value)

		else:
			#calls url page, changes page by cycling gene numbers
			Gene_Number = sheet.cell(row=i+1, column=1).value
			url = 'https://www.ncbi.nlm.nih.gov/gene/%s?report=expression' %Gene_Number
			res = requests.get(url)
			soup = bs(res.text, 'lxml')
			#second webpage to get geneloc, exon count, chrom num, first and last gene number
			url_2 = 'https://www.ncbi.nlm.nih.gov/gene/%s' %Gene_Number
			res_2 = requests.get(url_2)
			soup_2 = bs(res_2.text, 'lxml')
			tree = html.fromstring(res_2.content)

			#--------------------------------------
			#this code is using url_2
			location = tree.xpath('/html/body/div[1]/div[1]/form/div[1]/div[4]/div/div[6]/div[2]/div[2]/div[2]/div/div[1]/dl/dd/span/text()')
			exon_count = tree.xpath('/html/body/div[1]/div[1]/form/div[1]/div[4]/div/div[6]/div[2]/div[2]/div[2]/div/dl/dd/text()')

			table = soup_2.find('table', attrs={'class': 'jig-ncbigrid'})

			rows = table.findAll('td')

			both_locs = rows[4].get_text()
			char1 = '('
			char2 = ')'
			bot_loc_with_periods = both_locs[both_locs.find(char1) + 1 : both_locs.find(char2)]
			both_loc_list = bot_loc_with_periods.split('..')

			#----------------------

			#var script is the coding off webpage that contains the javascript
			#which contains the variables we want.  This returns all the scripts
			#in a list

			script = soup.find_all('script', attrs={'type': 'text/javascript'})

			#want script[5], the 6th script contains all the variable.  Isolate.
			want = script[5]
			#print(want)
			#convert from soup obj to string; removes html code
			str_want = want.string
			#print('length:', len(str_want))

			#skips pages that have no rpkm values if the page isn't coded with javascript
			if str_want == None:
				error_list.append(i+1)
				sheet.cell(row=i+1, column=32).value = 'Error'
				print('error: ', i+1)
			#skips page if it is coded with js but has no data entered because the string will only be ~36 char
			elif len(str_want) < 500:
				error_list.append(i+1)
				sheet.cell(row=i+1, column=32).value = 'Error'
				print('error: ', i+1)

			#Data is stored in a weird file format.  Its multiple var =  {}; var = {};
			#this breaks it at first ';' allowing for first variable to be edited
			else:
				head, sep, tail = str_want.partition(';')
				#print(head)
				#print(len(head)) #lengh of bad head is 36
				#json files use double quotes.  These files are stored as single quotes.
				#I'm convinced its done to annoy me.  This fxn converts the single
				#quotes to double quotes. Then stripes the 'var =', leaving only
				#the dictionary like json values.  Json.load converts to python dict
				if len(head) < 50:
					error_list.append(i+1)
					sheet.cell(row=i+1, column=32).value = 'Error'
					print('error: ', i+1)

				else:

					head = head.replace("'", '"')
					#print(head)
					jsonValue = '{%s}' % (head.split('{', 1)[1].rsplit('}', 1)[0],)
					value = json.loads(jsonValue)

					nav_gene_name = soup.find('span', attrs={'class': 'gn'})

					#Getting each value from the dictionary is a pain since it appears to be a dict
					#embedded in a dict.  
					#anyways, this extracts each dictionary to a new variable correpsonding to its type
					#I don't think this step is actually necessary but I'm to tired to find a proper way to do it.
					#this line of code was creaded using tissue_names.py
					gene_name = nav_gene_name.string
					genomic_location = location[0]
					exon_count = exon_count[0]
					chromosome_number = rows[3].get_text()
					genomic_location_number_beginning = both_loc_list[0]
					genomic_location_number_ending = both_loc_list[1]

					liver  =  value['liver']
					gall_bladder  =  value['gall bladder']
					small_intestine  =  value['small intestine']
					heart  =  value['heart']
					stomach  =  value['stomach']
					kidney  =  value['kidney']
					pancreas  =  value['pancreas']
					endometrium  =  value['endometrium']
					placenta  =  value['placenta']
					spleen  =  value['spleen']
					brain  =  value['brain']
					prostate  =  value['prostate']
					appendix  =  value['appendix']
					lymph_node  =  value['lymph node']
					duodenum  =  value['duodenum']
					skin  =  value['skin']
					salivary_gland  =  value['salivary gland']
					ovary  =  value['ovary']
					lung  =  value['lung']
					testis  =  value['testis']
					colon  =  value['colon']
					urinary_bladder  =  value['urinary bladder']
					adrenal  =  value['adrenal']
					thyroid  =  value['thyroid']
					bone_marrow  =  value['bone marrow']
					fat  =  value['fat']
					esophagus  =  value['esophagus']

					#put dicts into a list.  I know its oxymoronic at this point
					tissue_list_var = [adrenal, appendix, bone_marrow, brain, colon, duodenum, endometrium, esophagus, fat, gall_bladder, heart, kidney, liver, lung, lymph_node, ovary, pancreas, placenta, prostate, salivary_gland, skin, small_intestine, spleen, stomach, testis, thyroid, urinary_bladder]
					c = 9 #start printing in excel at column 9

					#prints gene name into excel sheet
					sheet.cell(row=i+1, column=3).value = gene_name
					sheet.cell(row=i+1, column=4).value = genomic_location
					sheet.cell(row=i+1, column=5).value = exon_count
					sheet.cell(row=i+1, column=6).value = chromosome_number
					sheet.cell(row=i+1, column=7).value = genomic_location_number_beginning
					sheet.cell(row=i+1, column=8).value = genomic_location_number_ending
					#prints rpkm value for each tissue for each gene into excel sheet
					for name in tissue_list_var:
						sheet.cell(row=i+1, column=c).value = name['full_rpkm']
						c += 1

					# print('GeneID = ', sheet.cell(row=i+1, column=1).value)
					# print('testis rpkm =', testis['full_rpkm'])
					print('row', i+1, 'added,', 'cell num:', sheet.cell(row=i+1, column=1).value, ',', 'cell name:', sheet.cell(row=i+1, column=3).value)
					#There is a crash but that if it crashes sometimes the excel
					#sheet becomes corrupt. Maybe saving to a different sheet besides 
					#the original will save one of them from corrupting
					#or may corrupt both
					wb.save('data_match.xlsx')

	else:

		break
print(error_list)
wb.save('data_match.xlsx')
